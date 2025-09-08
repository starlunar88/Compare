from flask import Flask, render_template, request, jsonify
import PyPDF2
import os
import difflib
import time
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB 제한
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')

# 배포 환경에서 포트 설정
port = int(os.environ.get('PORT', 5000))

# 업로드 폴더 생성
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def safe_remove_file(file_path, max_retries=3):
    """파일을 안전하게 삭제하는 함수"""
    for attempt in range(max_retries):
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                return True
        except OSError as e:
            if attempt < max_retries - 1:
                time.sleep(0.1)  # 잠시 대기 후 재시도
                continue
            else:
                print(f"파일 삭제 실패 (시도 {max_retries}회): {file_path} - {e}")
                return False
    return False

def cleanup_uploads_folder():
    """업로드 폴더의 오래된 파일들을 정리하는 함수"""
    try:
        uploads_dir = app.config['UPLOAD_FOLDER']
        if os.path.exists(uploads_dir):
            for filename in os.listdir(uploads_dir):
                file_path = os.path.join(uploads_dir, filename)
                if os.path.isfile(file_path):
                    # 파일이 1시간 이상 오래된 경우 삭제
                    if time.time() - os.path.getmtime(file_path) > 3600:
                        safe_remove_file(file_path)
    except Exception as e:
        print(f"업로드 폴더 정리 중 오류: {e}")

def extract_pdf_text(pdf_path):
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        return f"PDF 읽기 오류: {str(e)}"

def extract_excel_data(excel_path):
    try:
        print(f"엑셀 파일 경로: {excel_path}")
        print(f"파일 존재 여부: {os.path.exists(excel_path)}")
        
        # openpyxl로 직접 읽기
        from openpyxl import load_workbook
        
        all_values = []
        
        try:
            # 방법 1: openpyxl로 직접 읽기
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            print(f"워크시트 활성화됨: {ws.title}")
            
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if value is not None and str(value).strip():
                        all_values.append(str(value).strip())
                        
        except Exception as e1:
            print(f"방법1 실패: {e1}")
            
            try:
                # 방법 2: 모든 시트 시도
                wb = load_workbook(excel_path, data_only=True)
                print(f"모든 시트: {wb.sheetnames}")
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"시트 '{sheet_name}' 처리 중...")
                    
                    for row in ws.iter_rows(values_only=True):
                        for value in row:
                            if value is not None and str(value).strip():
                                all_values.append(str(value).strip())
                                
            except Exception as e2:
                print(f"방법2 실패: {e2}")
        
        print(f"추출된 모든 값: {all_values}")
        
        if not all_values:
            return {"error": "엑셀 파일에서 데이터를 찾을 수 없습니다."}
        
        # 간단한 형태로 반환
        result = {
            "Sheet1": [{"value": val} for val in all_values]
        }
        
        print(f"최종 엑셀 데이터: {result}")
        return result
        
    except Exception as e:
        print(f"엑셀 읽기 오류: {str(e)}")
        return {"error": f"엑셀 읽기 오류: {str(e)}"}

def compare_text_content(pdf_text, excel_data):
    differences = []
    
    excel_text = ""
    for sheet_name, data in excel_data.items():
        if isinstance(data, dict) and "error" in data:
            differences.append({"type": "error", "message": data["error"]})
            return differences
            
        # 시트 이름과 구조 정보 없이 순수 데이터만 추출
        for row in data:
            for key, value in row.items():
                if value and str(value).strip():  # 빈 값이 아닌 경우만
                    excel_text += f"{value} "
    
    pdf_lines = pdf_text.split('\n')
    excel_lines = excel_text.strip().split()
    
    diff = list(difflib.unified_diff(pdf_lines, excel_lines, fromfile='PDF', tofile='Excel', lineterm=''))
    
    if diff:
        differences.append({"type": "text_diff", "content": diff})
    
    # 공백과 특수문자 정리 후 단어 비교
    import re
    
    pdf_clean = re.sub(r'[^\w\s]', ' ', pdf_text.lower())
    excel_clean = re.sub(r'[^\w\s]', ' ', excel_text.lower())
    
    pdf_words = set(pdf_clean.split())
    excel_words = set(excel_clean.split())
    
    # 빈 문자열과 숫자만 있는 단어 제외
    pdf_words = {word for word in pdf_words if word and len(word) > 0}
    excel_words = {word for word in excel_words if word and len(word) > 0}
    
    only_in_pdf = pdf_words - excel_words
    if only_in_pdf:
        differences.append({"type": "only_in_pdf", "content": list(only_in_pdf)})
    
    only_in_excel = excel_words - pdf_words
    if only_in_excel:
        differences.append({"type": "only_in_excel", "content": list(only_in_excel)})
    
    return differences

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        # 업로드 폴더 정리
        cleanup_uploads_folder()
        
        if 'pdf_file' not in request.files or 'excel_file' not in request.files:
            return jsonify({"error": "PDF와 엑셀 파일을 모두 업로드해주세요."}), 400
        
        pdf_file = request.files['pdf_file']
        excel_file = request.files['excel_file']
        
        if pdf_file.filename == '' or excel_file.filename == '':
            return jsonify({"error": "파일을 선택해주세요."}), 400
        
        if not (allowed_file(pdf_file.filename) and allowed_file(excel_file.filename)):
            return jsonify({"error": "지원하지 않는 파일 형식입니다."}), 400
        
        pdf_filename = secure_filename(pdf_file.filename)
        excel_filename = secure_filename(excel_file.filename)
        
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        
        pdf_file.save(pdf_path)
        excel_file.save(excel_path)
        
        pdf_text = extract_pdf_text(pdf_path)
        excel_data = extract_excel_data(excel_path)
        differences = compare_text_content(pdf_text, excel_data)
        
        # 파일 삭제 시 안전한 오류 처리
        safe_remove_file(pdf_path)
        safe_remove_file(excel_path)
        
        return jsonify({
            "success": True,
            "pdf_text": pdf_text[:1000] + "..." if len(pdf_text) > 1000 else pdf_text,
            "excel_data": excel_data,
            "differences": differences
        })
        
    except Exception as e:
        return jsonify({"error": f"처리 중 오류가 발생했습니다: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=port)